"""Excel 读写工具（基于 openpyxl）

提供通用的 xlsx 构建与解析：
- build_xlsx：将表头 + 行数据写入内存中的 xlsx，返回字节
- parse_xlsx：解析 xlsx 字节为字典列表。当提供 expected_headers 时，会自动在
  所有工作表中、并扫描每个表的前若干行，定位包含全部必需列的表头行
  （容忍多工作表、表头上方有标题行等情况）。
"""
from io import BytesIO
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook

# 寻找表头时，扫描每个工作表的前 N 行
HEADER_SCAN_ROWS = 15


class ExcelParseError(Exception):
    """Excel 解析错误"""
    pass


def build_xlsx(headers: List[str], rows: List[List[Any]],
               sheet_title: str = "Sheet1") -> bytes:
    """构建 xlsx 并返回字节

    Args:
        headers: 表头列名
        rows: 行数据，每行为与 headers 等长的值列表
        sheet_title: 工作表名
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _row_to_strs(row) -> List[str]:
    return [str(c).strip() if c is not None else "" for c in (row or [])]


def _is_blank(row) -> bool:
    return row is None or all(c is None or str(c).strip() == "" for c in row)


def _parse_sheet(ws, expected_headers: Optional[List[str]]) -> Optional[List[Dict[str, Any]]]:
    """解析单个工作表。

    - 无 expected_headers：以第一行为表头。
    - 有 expected_headers：在前 HEADER_SCAN_ROWS 行中查找包含全部必需列的表头行；
      找不到则返回 None（表示此表不匹配）。
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    if expected_headers:
        header_idx = None
        for i in range(min(HEADER_SCAN_ROWS, len(rows))):
            hs = _row_to_strs(rows[i])
            if all(h in hs for h in expected_headers):
                header_idx = i
                break
        if header_idx is None:
            return None
    else:
        header_idx = 0

    headers = _row_to_strs(rows[header_idx])

    result: List[Dict[str, Any]] = []
    for row in rows[header_idx + 1:]:
        if _is_blank(row):
            continue
        record: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) else None
        result.append(record)
    return result


def parse_xlsx(data: bytes, expected_headers: Optional[List[str]] = None,
               sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """解析 xlsx 字节为字典列表。

    Args:
        data: xlsx 文件字节
        expected_headers: 若提供，会在所有工作表中查找包含这些列的表头行
        sheet_name: 指定工作表名（提供时只在该表内查找表头）

    Returns:
        每行一个 dict，键为表头，值为单元格值。空行被跳过。
    """
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise ExcelParseError(f"Failed to read xlsx: {e}")

    try:
        if sheet_name:
            res = _parse_sheet(wb[sheet_name], expected_headers)
            if res is None and expected_headers:
                raise ExcelParseError(f"Missing required columns: {expected_headers}")
            return res or []

        if expected_headers:
            # 在所有工作表中查找含必需列的表头
            for ws in wb.worksheets:
                res = _parse_sheet(ws, expected_headers)
                if res is not None:
                    return res
            raise ExcelParseError(
                f"Missing required columns: {expected_headers}（已扫描全部工作表均未找到）"
            )

        # 无必需列要求：活动表第一行作表头
        return _parse_sheet(wb.active, None) or []
    finally:
        wb.close()
